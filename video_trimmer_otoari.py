import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ffmpeg
import threading
import os
import json
import numpy as np
from scipy.io import wavfile
from scipy.fft import fft
import openpyxl
from datetime import datetime, timedelta
import traceback  # スタックトレース出力用
import sys  # システムエラー出力用

class VideoTrimmerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("動画トリマー")
        self.root.geometry("600x300")
        
        # 変数の初期化
        self.input_path = tk.StringVar()
        self.start_time = tk.StringVar(value="0.0")
        self.end_time = tk.StringVar()
        
        self.setup_ui()
    
    def setup_ui(self):
        # 入力ファイル選択
        input_frame = ttk.LabelFrame(self.root, text="入力ファイル(動画･音声)", padding=10)
        input_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Entry(input_frame, textvariable=self.input_path, width=50).pack(side="left", padx=5)
        ttk.Button(input_frame, text="参照...", command=self.select_input_file).pack(side="left")
        
        # 時間設定
        time_frame = ttk.LabelFrame(self.root, text="時間設定", padding=10)
        time_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(time_frame, text="開始時間（秒）:").pack(side="left")
        ttk.Entry(time_frame, textvariable=self.start_time, width=10).pack(side="left", padx=5)
        
        ttk.Label(time_frame, text="終了時間（秒）:").pack(side="left")
        ttk.Entry(time_frame, textvariable=self.end_time, width=10).pack(side="left", padx=5)
        
        # 動画情報表示
        self.info_label = ttk.Label(self.root, text="")
        self.info_label.pack(pady=5)
        
        # 出力ファイル情報
        self.output_label = ttk.Label(self.root, text="")
        self.output_label.pack(pady=5)
        
        # ボタンフレーム
        button_frame = ttk.Frame(self.root)
        button_frame.pack(pady=10)
        
        # トリミングボタン
        ttk.Button(button_frame, text="トリミング実行", command=self.trim_video).pack(side="left", padx=5)
        
        # 音声抽出ボタン
        ttk.Button(button_frame, text="音声抽出", command=self.extract_audio).pack(side="left", padx=5)
        
        # フーリエ変換ボタン
        ttk.Button(button_frame, text="フーリエ変換", command=self.analyze_audio).pack(side="left", padx=5)
        
        # プログレスバー
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.root, 
            variable=self.progress_var, 
            maximum=100
        )
        self.progress_bar.pack(fill="x", padx=10, pady=5)
    
    def select_input_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("メディアファイル", "*.mp4 *.avi *.mov *.mp3 *.wav"),
                ("動画ファイル", "*.mp4 *.avi *.mov"),
                ("音声ファイル", "*.mp3 *.wav"),
                ("すべてのファイル", "*.*")
            ]
        )
        if file_path:
            if os.path.exists(file_path):
                self.input_path.set(file_path)
                self.load_video_info()
                self.update_output_path()
            else:
                messagebox.showerror("エラー", f"指定したファイルが見つかりません：\n{file_path}")
    
    def generate_output_path(self, input_path, suffix="_trimming", ext=".mp4"):
        directory = os.path.dirname(input_path)
        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        
        # 新しいファイル名のベース
        output_base = os.path.join(directory, f"{name}{suffix}")
        
        # 指定された拡張子を使用
        counter = 1
        while True:
            output_path = f"{output_base}{counter:02d}{ext}"
            if not os.path.exists(output_path):
                return output_path
            counter += 1
    
    def update_output_path(self):
        if self.input_path.get():
            output_path = self.generate_output_path(self.input_path.get())
            # 拡張子を除いたファイル名を表示
            filename = os.path.basename(output_path)
            filename_without_ext = os.path.splitext(filename)[0]
            self.output_label.config(text=f"出力ファイル: {filename_without_ext}")
    
    def load_video_info(self):
        input_path = self.input_path.get()
        if not os.path.exists(input_path):
            messagebox.showerror("エラー", f"ファイルが見つかりません：\n{input_path}")
            return
        
        try:
            # FFmpegでファイル情報を取得
            probe = ffmpeg.probe(input_path)
            duration = float(probe['format']['duration'])
            
            # ファイルサイズを取得（MB単位）
            file_size = os.path.getsize(input_path) / (1024 * 1024)
            
            # 動画ファイルの場合は追加情報を表示
            if any(s['codec_type'] == 'video' for s in probe['streams']):
                video_info = next(s for s in probe['streams'] if s['codec_type'] == 'video')
                width = int(video_info['width'])
                height = int(video_info['height'])
                fps = eval(video_info.get('r_frame_rate', '30'))
                info_text = f"動画情報: {width}x{height}, {fps:.1f}fps, 長さ: {duration:.1f}秒, サイズ: {file_size:.1f}MB"
            else:
                info_text = f"音声情報: 長さ: {duration:.1f}秒, サイズ: {file_size:.1f}MB"
            self.info_label.config(text=info_text)
            
            # 終了時間を動画の長さに設定
            self.end_time.set(f"{duration:.1f}")
            
        except Exception as e:
            messagebox.showerror("エラー", f"動画ファイルを開けません：\n{input_path}\n{str(e)}")
    
    def trim_video(self):
        input_path = self.input_path.get()
        if not input_path:
            messagebox.showerror("エラー", "入力動画を選択してください。")
            return
        
        if not os.path.exists(input_path):
            messagebox.showerror("エラー", f"入力ファイルが見つかりません：\n{input_path}")
            return
        
        try:
            start_time = float(self.start_time.get())
            end_time = float(self.end_time.get())
        except ValueError:
            messagebox.showerror("エラー", "開始時間と終了時間は数値で入力してください。")
            return
        
        if start_time >= end_time:
            messagebox.showerror("エラー", "開始時間は終了時間より前である必要があります。")
            return
        
        # トリミング処理を別スレッドで実行
        threading.Thread(target=self.trim_video_thread).start()
    
    def trim_video_thread(self):
        try:
            start_time = float(self.start_time.get())
            end_time = float(self.end_time.get())
            input_path = self.input_path.get()
            output_path = self.generate_output_path(input_path)
            duration = end_time - start_time
            
            # FFmpegストリームを設定
            stream = ffmpeg.input(input_path, ss=start_time, t=duration)
            
            # 音声と映像を同時に処理
            stream = ffmpeg.output(stream, output_path,
                                 acodec='aac',       # 音声コーデック
                                 vcodec='libx264',   # 映像コーデック
                                 loglevel='error')   # エラーのみ表示
            
            # FFmpegコマンドを実行
            ffmpeg.run(stream, capture_stdout=True, capture_stderr=True, overwrite_output=True)
            
            if os.path.exists(output_path):
                self.root.after(0, self.trim_completed, output_path)
            else:
                raise Exception("出力ファイルが生成されませんでした。")
            
        except ffmpeg.Error as e:
            error_msg = f"FFmpegエラー：\n{e.stderr.decode()}"
            print(f"[ERROR] {error_msg}")  # ターミナルに出力
            self.root.after(0, self.show_error, error_msg)
        except Exception as e:
            error_msg = f"エラーが発生しました：\n{str(e)}"
            print(f"[ERROR] {error_msg}")  # ターミナルに出力
            self.root.after(0, self.show_error, error_msg)
    
    def show_error(self, message):
        self.progress_var.set(0)
        messagebox.showerror("エラー", message)
    
    def trim_completed(self, output_path):
        self.progress_var.set(0)
        # 出力ファイルのサイズを取得（MB単位）
        file_size = os.path.getsize(output_path) / (1024 * 1024)
        messagebox.showinfo("完了", f"動画のトリミングが完了しました。\n保存先: {output_path}\nファイルサイズ: {file_size:.1f}MB")
        # 出力パスの表示を更新
        self.update_output_path()

    def extract_audio(self):
        input_path = self.input_path.get()
        if not input_path:
            messagebox.showerror("エラー", "入力動画を選択してください。")
            return
        
        if not os.path.exists(input_path):
            messagebox.showerror("エラー", f"入力ファイルが見つかりません：\n{input_path}")
            return
        
        # 音声抽出処理を別スレッドで実行
        threading.Thread(target=self.extract_audio_thread).start()
    
    def extract_audio_thread(self):
        try:
            input_path = self.input_path.get()
            output_path = self.generate_output_path(input_path, suffix="_audio", ext=".mp3")
            
            # FFmpegストリームを設定
            stream = ffmpeg.input(input_path)
            
            # 音声のみを抽出してMP3として保存
            stream = ffmpeg.output(stream, output_path,
                               acodec='libmp3lame',  # MP3コーデック
                               vn=None,              # 映像を除外
                               loglevel='error')     # エラーのみ表示
            
            # FFmpegコマンドを実行
            ffmpeg.run(stream, capture_stdout=True, capture_stderr=True, overwrite_output=True)
            
            if os.path.exists(output_path):
                self.root.after(0, self.audio_extraction_completed, output_path)
            else:
                raise Exception("出力ファイルが生成されませんでした。")
            
        except ffmpeg.Error as e:
            error_msg = f"FFmpegエラー：\n{e.stderr.decode()}"
            print(f"[ERROR] {error_msg}")  # ターミナルに出力
            self.root.after(0, self.show_error, error_msg)
        except Exception as e:
            error_msg = f"エラーが発生しました：\n{str(e)}"
            print(f"[ERROR] {error_msg}")  # ターミナルに出力
            self.root.after(0, self.show_error, error_msg)
    
    def audio_extraction_completed(self, output_path):
        self.progress_var.set(0)
        # 出力ファイルのサイズを取得（MB単位）
        file_size = os.path.getsize(output_path) / (1024 * 1024)
        messagebox.showinfo("完了", f"音声抽出が完了しました。\n保存先: {output_path}\nファイルサイズ: {file_size:.1f}MB")
        # 出力パスの表示を更新
        self.update_output_path()

    def frequency_to_note(self, freq):
        """周波数を音階に変換（国際式とドレミ式）"""
        if freq == 0:
            return "無音", "無音"
        
        # A4(440Hz)を基準に周波数から音階を計算
        notes_international = ['C', 'C#', 'D', 'D#', 'E', 'F', 'F#', 'G', 'G#', 'A', 'A#', 'B']
        notes_doremi = ['ド', 'ド#', 'レ', 'レ#', 'ミ', 'ファ', 'ファ#', 'ソ', 'ソ#', 'ラ', 'ラ#', 'シ']
        
        a4_freq = 440
        steps = round(12 * np.log2(freq / a4_freq))
        octave = 4 + (steps + 9) // 12
        note_index = (steps + 9) % 12
        
        # 国際式とドレミ式の音名を生成
        note_international = f"{notes_international[note_index]}{octave}"
        note_doremi = f"{notes_doremi[note_index]}{octave}"
        
        return note_international, note_doremi

    def analyze_audio(self):
        input_path = self.input_path.get()
        if not input_path:
            messagebox.showerror("エラー", "入力ファイルを選択してください。")
            return
        
        if not os.path.exists(input_path):
            messagebox.showerror("エラー", f"入力ファイルが見つかりません：\n{input_path}")
            return
        
        # 音声解析処理を別スレッドで実行
        threading.Thread(target=self.analyze_audio_thread).start()

    def analyze_audio_thread(self):
        try:
            input_path = self.input_path.get()
            
            # 入力がMP3の場合、一時的なWAVファイルに変換
            if input_path.lower().endswith('.mp3'):
                wav_path = input_path.rsplit('.', 1)[0] + '_temp.wav'
                stream = ffmpeg.input(input_path)
                stream = ffmpeg.output(stream, wav_path, acodec='pcm_s16le', ac=1)
                ffmpeg.run(stream, capture_stdout=True, capture_stderr=True, overwrite_output=True)
                input_path = wav_path
            
            # WAVファイルを読み込み
            sample_rate, audio_data = wavfile.read(input_path)
            if len(audio_data.shape) > 1:
                audio_data = audio_data[:, 0]  # ステレオの場合は最初のチャンネルを使用
            
            # 分析のパラメータ
            window_size = int(0.05 * sample_rate)  # 50ms
            hop_size = int(0.025 * sample_rate)    # 25ms
            
            # Excelファイルを作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "音声解析結果"
            
            # ヘッダーを設定
            headers = ["時刻(hh:mm:ss:fff)", "周波数 (Hz)", "振幅 (dB)", "音階（国際式）", "音階（ドレミ式）"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # データ解析と書き込み
            row = 2
            total_windows = (len(audio_data) - window_size) // hop_size
            
            for i in range(0, len(audio_data) - window_size, hop_size):
                # 進捗更新
                progress = (i / (len(audio_data) - window_size)) * 100
                self.root.after(0, self.progress_var.set, progress)
                
                # 時間窓でのデータを取得
                window = audio_data[i:i + window_size]
                
                # フーリエ変換
                spectrum = fft(window)
                freq = np.fft.fftfreq(window_size, 1/sample_rate)
                
                # 正の周波数のみを使用
                pos_mask = freq > 0
                freq = freq[pos_mask]
                spectrum = np.abs(spectrum[pos_mask])
                
                # 時刻を計算（ミリ秒まで）
                time_sec = i / sample_rate
                total_ms = int(time_sec * 1000)
                hours = total_ms // 3600000
                minutes = (total_ms % 3600000) // 60000
                seconds = (total_ms % 60000) // 1000
                milliseconds = total_ms % 1000
                time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}:{milliseconds:03d}"
                
                # 最大振幅とその周波数を特定
                max_idx = np.argmax(spectrum)
                max_amplitude = spectrum[max_idx]
                
                # データをExcelに書き込み（振幅が0の場合は周波数と音階を空欄に）
                ws.cell(row=row, column=1, value=time_str)  # 時刻は常に記録
                ws.cell(row=row, column=3, value=f"{max_amplitude:.1f}")  # 振幅は常に記録
                
                if max_amplitude > 0:  # 振幅が0より大きい場合のみ周波数と音階を記録
                    max_freq = freq[max_idx]
                    note_international, note_doremi = self.frequency_to_note(max_freq)
                    ws.cell(row=row, column=2, value=f"{max_freq:.1f}")
                    ws.cell(row=row, column=4, value=note_international)
                    ws.cell(row=row, column=5, value=note_doremi)
                
                row += 1
            
            # Excelファイルを保存
            excel_path = self.generate_output_path(self.input_path.get(), suffix="_analysis", ext=".xlsx")
            wb.save(excel_path)
            
            # 一時ファイルの削除
            if input_path.endswith('_temp.wav'):
                os.remove(input_path)
            
            self.root.after(0, self.analysis_completed, excel_path)
            
        except Exception as e:
            error_msg = f"解析エラー：\n{str(e)}\nスタックトレース:\n{traceback.format_exc()}"
            print(f"[ERROR] {error_msg}")  # ターミナルに出力
            self.root.after(0, self.show_error, error_msg)
    
    def analysis_completed(self, output_path):
        self.progress_var.set(0)
        messagebox.showinfo("完了", f"音声解析が完了しました。\n保存先: {output_path}")
        self.update_output_path()

if __name__ == "__main__":
    root = tk.Tk()
    app = VideoTrimmerGUI(root)
    root.mainloop()